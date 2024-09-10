import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import time

class FormatExcel:
    
    def __init__(self, filepath, output_file_path=None):
        self.filepath = filepath
        self.output_file_path = output_file_path
        self.workbook = load_workbook(self.filepath)
        self.worksheet = self.workbook.active
        self.change_colors()
        self.change_font()
        self.set_column_width()  # Call the method to set column width
        self.save_output_file()  # Save the workbook after all modifications

    def save_output_file(self):

        if self.output_file_path is None:
            self.output_file_path = os.path.splitext(self.filepath)[0] + "-hhoutput.xlsx"
            print(f"file name is: {os.path.join(self.filepath)[0]}")
        if os.path.exists(self.output_file_path):
            os.remove(self.output_file_path)

        output_dir = os.path.dirname(self.output_file_path)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            os.makedirs(output_dir+'.ipynb')
        self.workbook.save(self.output_file_path)
        print("saved at ", self.output_file_path)
        #self.workbook.close()  # Close the workbook after saving

    def delete_original_file(self):
        time.sleep(1)
        if os.path.exists(self.filepath):
            os.remove(self.filepath)

    def change_colors(self):
        fill = PatternFill(start_color='ff4300',
                           end_color='ff4300',
                           fill_type='solid')
        
        fill_first_row = PatternFill(start_color='ffff00',
                                     end_color='ffff00',
                                     fill_type='solid')

        # Get the column index of "Amount"
        amount_col_index = None
        for col in self.worksheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == "Difference":
                    amount_col_index = cell.column
                    break
            if amount_col_index is not None:
                break

        # If "Amount" column exists, change color based on condition
        if amount_col_index is not None:
            for row in range(2, self.worksheet.max_row + 1):
                amount_cell = self.worksheet.cell(row=row, column=amount_col_index)
                if amount_cell.value != 0:
                    amount_cell.fill = fill

        # Change color of first row
        for col in range(1, self.worksheet.max_column + 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.fill = fill_first_row

    def change_font(self):
        # Specify the font style you want to apply
        font = Font(name='Times New Roman', size=15, bold=True, italic=False, color='000080')
        
        # Apply the font style to the first row
        for col in range(1, self.worksheet.max_column + 1):
            cell = self.worksheet.cell(row=1, column=col)
            cell.font = font
    
    def set_column_width(self):
        # Set width for the first 5 columns
        for col in range(1, 23):
            self.worksheet.column_dimensions[chr(64 + col)].width = 20  # Assuming you want a width of 20 for each column

# Example usage:
    

